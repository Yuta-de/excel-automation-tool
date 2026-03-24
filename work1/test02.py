import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers

# 入力フォルダ
input_folder = r"C:\work\python_study\excel_automation\work1\売上データ元\*.xlsx"

# 読み込んだデータを入れるリスト
df_list = []

# フォルダ内のExcelをすべて取得
for file in glob.glob(input_folder):
    print("読み込み中", file)
    df = pd.read_excel(file)

    # 店舗名をファイル名から取得して列に追加
    store_name = os.path.basename(file).replace(".xlsx","")
    df["店舗名"] = store_name

    df_list.append(df)

# すべて結合
merged_df = pd.concat(df_list, ignore_index=True)

print("\n=== 結合結果_before ===")
print(merged_df)

# ======== 整形処理ここから ========
# 1. 日付を日付型に変換
merged_df["日付"] = pd.to_datetime(merged_df["日付"], errors="coerce")

# 2. 数値列を数値型に変換
num_cols = ["数量", "単価", "売上"]
for col in num_cols:
    merged_df[col] = pd.to_numeric(merged_df[col], errors="coerce")

# 3. 列の並び替え
desired_order = ["店舗名", "日付", "商品名", "カテゴリ", "数量", "単価", "売上"]
merged_df = merged_df[desired_order]

# 4. 日付でソート
merged_df = merged_df.sort_values(by=["日付", "店舗名"]).reset_index(drop=True)
# ======== 整形処理ここまで ========

print("\n=== 結合結果_after ===")
print(merged_df)

# 出力先ファイル
output_path = r"C:\work\python_study\excel_automation\work1\データ出力\merged_sales.xlsx"

# Excelに出力
merged_df.to_excel(output_path, index=False)

print("\n=== 出力完了 ===")
print("出力ファイル:", output_path)

# 書式設定のために openpyxl で読み込み
wb = load_workbook(output_path)
ws = wb.active

# 「日付」列の列番号を取得（1始まり）
data_col = list(merged_df.columns).index("日付") + 1

# 2行目以降に書式を適用
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=data_col)
    cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

wb.save(output_path)

print("日付の書式設定が完了しました:", output_path)