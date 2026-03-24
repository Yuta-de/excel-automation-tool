import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers

# 入力フォルダ
input_folder = r"C:\work\python_study\excel_automation\work1\売上データ元\*.xlsx"

# 読み込んだデータを入れるリスト
df_list = []

# フォルダ内のExcelをすべて取得
for file in glob.glob(input_folder):
    print("読み込み中", file)
    df = pd.read_excel(file)

    # 店舗名をファイル名から取得して列に追加
    store_name = os.path.basename(file).replace(".xlsx","")
    df["店舗名"] = store_name

    df_list.append(df)

# すべて結合
merged_df = pd.concat(df_list, ignore_index=True)

print("\n=== 結合結果_before ===")
print(merged_df)

# ======== 整形処理ここから ========
# 1. 日付を日付型に変換
merged_df["日付"] = pd.to_datetime(merged_df["日付"], errors="coerce")

# 2. 数値列を数値型に変換
num_cols = ["数量", "単価", "売上"]
for col in num_cols:
    merged_df[col] = pd.to_numeric(merged_df[col], errors="coerce")

# 3. 列の並び替え
desired_order = ["店舗名", "日付", "商品名", "カテゴリ", "数量", "単価", "売上"]
merged_df = merged_df[desired_order]

# 4. 日付でソート
merged_df = merged_df.sort_values(by=["日付", "店舗名"]).reset_index(drop=True)
# ======== 整形処理ここまで ========

print("\n=== 結合結果_after ===")
print(merged_df)

# ======== 集計処理ここから ========
# ① カテゴリ別売上集計
pivot_category = merged_df.pivot_table(
    index = "カテゴリ",
    values = "売上",
    aggfunc = "sum"
)
print("\n=== カテゴリ別売上 ===")
print(pivot_category)

# ② 商品別売上集計
pivot_item = merged_df.pivot_table(
    index = "商品名",
    values = "売上",
    aggfunc = "sum"
)
print("\n=== 商品別売上 ===")
print(pivot_item)

# ③ 店舗別売上集計
pivot_store = merged_df.pivot_table(
    index = "店舗名",
    values = "売上",
    aggfunc = "sum"
)
print("\n=== 店舗別売上 ===")
print(pivot_store)

# ④ 日付別売上推移
pivot_date = merged_df.pivot_table(
    index = "日付",
    values = "売上",
    aggfunc = "sum"
)
print("\n=== 日付別売上 ===")
print(pivot_date)
# ======== 集計処理ここまで ========

# 出力先ファイル
output_path = r"C:\work\python_study\excel_automation\work1\データ出力\merged_sales.xlsx"

# Excelに出力
merged_df.to_excel(output_path, index=False)

print("\n=== 出力完了 ===")
print("出力ファイル:", output_path)

# ======== 集計結果をシートに書き込む ========
wb = load_workbook(output_path)

def write_df_to_sheet(df, sheet_name):
    # シートが存在すれば削除して作り直す
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # ヘッダー行
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # データ行
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

# 4つの集計結果を書き込む
write_df_to_sheet(pivot_category.reset_index(), "カテゴリ別売上")
write_df_to_sheet(pivot_item.reset_index(), "商品別売上")
write_df_to_sheet(pivot_store.reset_index(), "店舗別売上")
write_df_to_sheet(pivot_date.reset_index(), "日付別売上")

# 保存
wb.save(output_path)

print("\n=== 集計結果をExcelに書き出しました ===")