import pandas as pd
import glob
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from openpyxl.chart.shapes import GraphicalProperties
from typing import Dict
from openpyxl.workbook import Workbook

# ==========================
# 1. データ読み込み
# ==========================
def load_sales_files(input_folder : str) -> pd.DataFrame:
    df_list = []
    for file in glob.glob(input_folder):
        print("読み込み中：", file)
        df = pd.read_excel(file)
        df["店舗名"] = os.path.basename(file).replace(".xlsx", "")
        df_list.append(df)
    return pd.concat(df_list, ignore_index=True)

# ==========================
# 2. 整形処理
# ==========================
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
    num_cols = ["数量", "単価", "売上"]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    
    desired_order = ["店舗名", "日付", "商品名", "カテゴリ", "数量", "単価", "売上"]
    df = df[desired_order]

    df = df.sort_values(by=["日付", "店舗名"]).reset_index(drop=True)
    return df
    
# ==========================
# 3. 集計処理
# ==========================
def create_pivots(df:pd.DataFrame) -> Dict[str, pd.DataFrame]:
    pivots = {
        "カテゴリ別売上": df.pivot_table(index="カテゴリ", values="売上", aggfunc="sum"),
        "商品別売上": df.pivot_table(index="商品名", values="売上", aggfunc="sum"),
        "店舗別売上": df.pivot_table(index="店舗名", values="売上", aggfunc="sum"),
        "日付別売上": df.pivot_table(index="日付", values="売上", aggfunc="sum")
    }
    return pivots

# ==========================
# 4. Excel書き込み
# ==========================
def write_df_to_sheet(wb:Workbook, df: pd.DataFrame, sheet_name: str) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

# ==========================
# 5. グラフ作成
# ==========================
def create_charts(wb: Workbook) -> None:
    # --- カテゴリ別（円グラフ） ---
    ws = wb["カテゴリ別売上"]
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    pie = PieChart()
    pie.title = "カテゴリ別売上"
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, "D2")

    # --- 店舗別（棒グラフ） ---
    ws = wb["店舗別売上"]
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    bar = BarChart()
    bar.title = "店舗別売上"
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.y_axis.title = "売上"
    bar.x_axis.title = "店舗名"
    bar.style = 10
    bar.width = 20
    bar.height = 10
    ws.add_chart(bar, "D2")

    # --- 商品別（横棒グラフ） ---
    ws = wb["商品別売上"]
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    bar2 = BarChart()
    bar2.title = "商品別売上"
    bar2.add_data(data, titles_from_data=True)
    bar2.set_categories(labels)
    bar2.y_axis.title = "売上"
    bar2.x_axis.title = "商品名"
    bar2.style = 10
    bar2.width = 20
    bar2.height = 10
    bar2.type = "bar"
    ws.add_chart(bar2, "D2")

    # --- 日付別（折れ線グラフ） ---
    ws = wb["日付別売上"]
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    line = LineChart()
    line.title = "日付別売上推移"
    line.add_data(data, titles_from_data=True)
    line.set_categories(labels)
    line.y_axis.title = "売上"
    line.x_axis.title = "日付"
    line.style = 13
    line.width = 20
    line.height = 12
    ws.add_chart(line, "D2")


# ==========================
# メイン処理
# ==========================
def main() -> None:
    input_folder = r"C:\work\python_study\excel_automation\work1\売上データ元\*.xlsx"
    output_path = r"C:\work\python_study\excel_automation\work1\データ出力\merged_sales.xlsx"

    df = load_sales_files(input_folder)
    df = clean_data(df)

    df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)

    pivots = create_pivots(df)
    for name, pivot_df in pivots.items():
        write_df_to_sheet(wb, pivot_df.reset_index(), name)
    
    create_charts(wb)
    wb.save(output_path)

    print("\n=== 完了しました！ ===")

main()