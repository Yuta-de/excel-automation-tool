import pandas as pd
import glob
import os
import logging
import configparser
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from openpyxl.chart.shapes import GraphicalProperties
from typing import Dict
from openpyxl.workbook import Workbook

if not os.path.exists("config.ini"):
    print("config.iniが見つかりません")
    exit()

# --- config.ini 読み込み ---
def load_config(path="config.ini") -> configparser.ConfigParser:
    config = configparser.ConfigParser()
    config.read(path, encoding="utf-8")
    return config

# ==========================
# 1. データ読み込み
# ==========================
def load_sales_files(input_folder : str) -> pd.DataFrame:
    logging.info("売上ファイルの読み込み開始")
    df_list = []
    for file in glob.glob(input_folder):
        try:
            logging.info(f"読み込み中： {file}")
            df = pd.read_excel(file)
            df["店舗名"] = os.path.basename(file).replace(".xlsx", "")
            df_list.append(df)
        except Exception as e:
            logging.error(f"読み込み失敗：{file} - {e}")
    logging.info("売上ファイルの読み込み完了")
    if not df_list:
        raise ValueError("Excelファイルが見つかりません")
    
    return pd.concat(df_list, ignore_index=True)

# ==========================
# 2. 整形処理
# ==========================
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    logging.info("整形処理開始")
    df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
    num_cols = ["数量", "単価", "売上"]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    
    desired_order = ["店舗名", "日付", "商品名", "カテゴリ", "数量", "単価", "売上"]
    df = df[desired_order]

    df = df.sort_values(by=["日付", "店舗名"]).reset_index(drop=True)

    logging.info("整形処理終了")
    return df
    
# ==========================
# 3. 集計処理
# ==========================
def create_pivots(df:pd.DataFrame) -> Dict[str, pd.DataFrame]:
    logging.info("集計処理開始")
    pivots = {
        "カテゴリ別売上": df.pivot_table(index="カテゴリ", values="売上", aggfunc="sum"),
        "商品別売上": df.pivot_table(index="商品名", values="売上", aggfunc="sum"),
        "店舗別売上": df.pivot_table(index="店舗名", values="売上", aggfunc="sum"),
        "日付別売上": df.pivot_table(index="日付", values="売上", aggfunc="sum")
    }
    logging.info("集計処理終了")
    return pivots

# ==========================
# 4. Excel書き込み
# ==========================
def write_df_to_sheet(wb:Workbook, df: pd.DataFrame, sheet_name: str) -> None:
    logging.info(f"シート書き込み： {sheet_name}")

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

# ==========================
# 5. グラフ作成
# ==========================
def create_charts(wb: Workbook) -> None:
    logging.info("グラフ作成開始")
    # --- カテゴリ別（円グラフ） ---
    logging.info("グラフ作成：カテゴリ別売上")
    ws = wb["カテゴリ別売上"]
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    pie = PieChart()
    pie.title = "カテゴリ別売上"
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, "D2")
    

    # --- 店舗別（棒グラフ） ---
    logging.info("グラフ作成：店舗別売上")
    ws = wb["店舗別売上"]
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    bar = BarChart()
    bar.title = "店舗別売上"
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.y_axis.title = "売上"
    bar.x_axis.title = "店舗名"
    bar.style = 10
    bar.width = 20
    bar.height = 10
    ws.add_chart(bar, "D2")
    

    # --- 商品別（横棒グラフ） ---
    logging.info("グラフ作成：商品別売上")
    ws = wb["商品別売上"]
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    bar2 = BarChart()
    bar2.title = "商品別売上"
    bar2.add_data(data, titles_from_data=True)
    bar2.set_categories(labels)
    bar2.y_axis.title = "売上"
    bar2.x_axis.title = "商品名"
    bar2.style = 10
    bar2.width = 20
    bar2.height = 10
    bar2.type = "bar"
    ws.add_chart(bar2, "D2")
    

    # --- 日付別（折れ線グラフ） ---
    logging.info("グラフ作成：日付別売上")
    ws = wb["日付別売上"]
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    line = LineChart()
    line.title = "日付別売上推移"
    line.add_data(data, titles_from_data=True)
    line.set_categories(labels)
    line.y_axis.title = "売上"
    line.x_axis.title = "日付"
    line.style = 13
    line.width = 20
    line.height = 12
    ws.add_chart(line, "D2")
    

    logging.info("全部のグラフ作成完了")

#コンフィグ読み込み
config = load_config()

# loggingの初期化
log_file = config["LOG"]["log_file"]
os.makedirs("logs", exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler()
    ]
)

# ==========================
# メイン処理
# ==========================
def main() -> None:
    try:
        logging.info("==== 処理開始 ====")

        input_folder = config["PATH"]["input_folder"]
        output_path = config["PATH"]["output_path"]

        df = load_sales_files(input_folder)
        df = clean_data(df)

        df.to_excel(output_path, index=False)
        logging.info("結合データのExcel出力完了")

        wb = load_workbook(output_path)

        pivots = create_pivots(df)
        for name, pivot_df in pivots.items():
            write_df_to_sheet(wb, pivot_df.reset_index(), name)
        
        create_charts(wb)
        wb.save(output_path)

        logging.info("=== 全処理完了 ===")
    except Exception as e:
        logging.error(f"致命的エラー発生：{e}")
        raise

    

main()