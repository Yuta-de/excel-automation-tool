import pandas as pd
import glob
import os
import logging
import configparser
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.chart import PieChart, Reference, BarChart, LineChart
from openpyxl.chart.shapes import GraphicalProperties
from typing import Dict
from openpyxl.workbook import Workbook

if not os.path.exists("config.ini"):
    print("config.iniが見つかりません")
    exit()

# --- config.ini 読み込み ---
def load_config(path="config.ini") -> configparser.ConfigParser:
    config = configparser.ConfigParser()
    config.read(path, encoding="utf-8")
    return config

# ==========================
# 1. データ読み込み
# ==========================
def load_sales_files(input_folder : str) -> pd.DataFrame:
    logging.info("売上ファイルの読み込み開始")
    df_list = []
    for file in glob.glob(input_folder):
        try:
            logging.info(f"読み込み中： {file}")
            df = pd.read_excel(file)
            df["店舗名"] = os.path.basename(file).replace(".xlsx", "")
            df_list.append(df)
        except Exception as e:
            logging.error(f"読み込み失敗：{file} - {e}")
    logging.info("売上ファイルの読み込み完了")
    if not df_list:
        raise ValueError("Excelファイルが見つかりません")
    
    return pd.concat(df_list, ignore_index=True)

# ==========================
# 2. 整形処理
# ==========================
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    logging.info("整形処理開始")
    df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
    num_cols = ["数量", "単価", "売上"]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    
    desired_order = ["店舗名", "日付", "商品名", "カテゴリ", "数量", "単価", "売上"]
    df = df[desired_order]

    df = df.sort_values(by=["日付", "店舗名"]).reset_index(drop=True)

    logging.info("整形処理終了")
    return df
    
# ==========================
# 3. 集計処理
# ==========================
def create_pivots(df:pd.DataFrame) -> Dict[str, pd.DataFrame]:
    logging.info("集計処理開始")
    pivots = {
        "カテゴリ別売上": df.pivot_table(index="カテゴリ", values="売上", aggfunc="sum"),
        "商品別売上": df.pivot_table(index="商品名", values="売上", aggfunc="sum"),
        "店舗別売上": df.pivot_table(index="店舗名", values="売上", aggfunc="sum"),
        "日付別売上": df.pivot_table(index="日付", values="売上", aggfunc="sum")
    }
    logging.info("集計処理終了")
    return pivots

# ==========================
# 4. Excel書き込み
# ==========================
def write_df_to_sheet(wb:Workbook, df: pd.DataFrame, sheet_name: str) -> None:
    logging.info(f"テンプレートシートへ書き込み： {sheet_name}")

    # 既存シートの取得
    ws = wb[sheet_name]

    # ヘッダー以外の表をクリア
    for row in ws["A2:Z999"]:
        for cell in row:
            cell.value = None

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

# ==========================
# 5. グラフ更新
# ==========================
def update_chart_ranges(wb:Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    ws = wb[sheet_name]
     # シート内の最初のグラフを取得

    max_row = len(df) + 1

    if not ws._charts: # type: ignore[attr-defined]
        logging.warning(f"{sheet_name}にグラフがありません")
        return
    chart = ws._charts[0] # type: ignore[attr-defined]

    chart.series[0].values = f"'{sheet_name}'!B2:B{max_row}"
    chart.series[0].categories = f"'{sheet_name}'!A2:A{max_row}"

    logging.info(f"グラフ更新完了：{sheet_name}")

#コンフィグ読み込み
config = load_config()

# loggingの初期化
log_file = config["LOG"]["log_file"]
os.makedirs("logs", exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler()
    ]
)

# ==========================
# メイン処理
# ==========================
def main() -> None:
    try:
        logging.info("==== 処理開始 ====")

        input_folder = config["PATH"]["input_folder"]
        output_path = config["PATH"]["output_path"]

        df = load_sales_files(input_folder)
        df = clean_data(df)

        template_path = config["PATH"]["template_path"]
        wb = load_workbook(template_path)

        # 統合データの書き込み
        write_df_to_sheet(wb, df, sheet_name="データ統合")

        pivots = create_pivots(df)
        for name, pivot_df in pivots.items():
            write_df_to_sheet(wb, pivot_df.reset_index(), name)
            update_chart_ranges(wb, name, pivot_df)
        
        wb.save(output_path)

        logging.info("=== 全処理完了 ===")
    except Exception as e:
        logging.error(f"致命的エラー発生：{e}")
        raise

main()